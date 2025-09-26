# streamlit_excel_cleaner.py
import os
import io
import pandas as pd
import streamlit as st
from io import BytesIO
import csv
from openpyxl.styles import PatternFill, Border, Side
from openpyxl import load_workbook

# List of columns to hide (delete)
columns_to_hide = [
    "Ride ID", "Pickup Time (UTC)", "Pickup Timezone offset from UTC", "Pickup Date (UTC)",
    "Drop-off Time (Local)", "Drop-off Time (UTC)", "Drop-off Timezone", "Drop-off Date (Local)", "Drop-off Date (UTC)", "Email",
    "Pickup City", "Pickup State", "Pickup Zip Code", "Requester Name",
    "Drop-off City", "Drop-off State", "Drop-off Zip Code",
    "Request Address", "Request City", "Request State", "Request Zip Code",
    "Destination Address", "Destination City", "Destination State", "Destination Zip Code",
    "Duration (minutes)", "Ride Fare", "Ride Fees", "Ride Discounts", "Ride Tip", "Ride Cost",
    "Business Services Fee", "Transaction Date (UTC)", "Transaction Time (UTC)", "Transaction Currency", "Transaction Outcome",
    "Expense Code", "Expense Note", "Ride Type", "Employee ID", "Custom Tag 1", "Custom Tag 2",
    "Fare Type", "Scheduled Ride Id", "Flex Ride Id", "Flex Ride", "Pickup Latitude", "Pickup Longitude",
    "Drop-off Latitude", "Drop-off Longitude",
    "Trip/Eats ID", "Transaction Timestamp (UTC)", "Request Date (UTC)", "Request Time (UTC)", "Request Date (Local)", "Request Time (Local)",
    "Request Type", "Request Timezone Offset from UTC", "Service", "City", "Haversine Distance (mi)", "Duration (min)", "Drop Off Latitude",
    "Drop Off Longitude", "Expense Code", "Invoices", "Program", "Group", "Payment Method", "Fare in Local Currency (excl. Taxes)", "Taxes in Local Currency",
    "Tip in Local Currency", "Taxes in Local Currency", "Tip in Local Currency",
    "Local Currency Code", "Fare in USD (excl. Taxes)", "Taxes in USD", "Tip in USD", "Transaction Amount in USD (incl. Taxes)", "Estimated Service and Technology Fee (incl. Taxes, if any) in USD",
    "Health Dashboard URL", "Invoice Number", "Driver First Name", "Deductions in Local Currency", "Member ID", "Plan ID", "Network Transaction Id",
    "IsGroupOrder", "Fulfilment Type", "Country", "Cancellation type", "Membership Savings(Local Currency)", "Granular Service Purpose Type"
]

# Define expected headers for uber file
expected_headers_uber = [
    "Trip/Eats ID", "Transaction Timestamp (UTC)", "Request Date (UTC)", "Request Time (UTC)", "Request Date (Local)", "Request Time (Local)", 
    "Request Type", "Pickup Date (UTC)", "Pickup Time (UTC)", "Pickup Date (Local)", "Pickup Time (Local)", 
    "Drop-off Date (UTC)", "Drop-off Time (UTC)", "Drop-off Date (Local)", "Drop-off Time (Local)", 
    "Request Timezone Offset from UTC", "First Name", "Last Name", "Email", "Employee ID", "Service", "City", 
    "Distance (mi)", "Haversine Distance (mi)", "Duration (min)", "Pickup Address", "Pickup Latitude", "Pickup Longitude", 
    "Drop-off Address", "Drop Off Latitude", "Drop Off Longitude", "Ride Status", "Expense Code", "Internal Note", 
    "Invoices", "Program", "Group", "Payment Method", "Transaction Type", 
    "Fare in Local Currency (excl. Taxes)", "Taxes in Local Currency", "Tip in Local Currency", 
    "Transaction Amount in Local Currency (incl. Taxes)", "Local Currency Code", 
    "Fare in USD (excl. Taxes)", "Taxes in USD", "Tip in USD", "Transaction Amount in USD (incl. Taxes)", 
    "Estimated Service and Technology Fee (incl. Taxes, if any) in USD", 
    "Health Dashboard URL", "Invoice Number", "Driver First Name", "Guest First Name", 
    "Guest Last Name", "Passenger Number", "Deductions in Local Currency", "Member ID", "Plan ID"
]

# Define expected headers for lyft file
expected_headers_lyft = [
    "Ride ID", "Pickup Date (UTC)", "Pickup Time (UTC)", "Pickup Date (Local)", "Pickup Time (Local)",
    "Pickup Timezone offset from UTC", "Drop-off Date (UTC)", "Drop-off Time (UTC)",
    "Drop-off Date (Local)", "Drop-off Time (Local)", "First Name", "Last Name", "Email",
    "Pickup Address", "Pickup City", "Pickup State", "Pickup Zip Code", "Drop-off Address",
    "Drop-off City", "Drop-off State", "Drop-off Zip Code", "Request Address", "Request City",
    "Request State", "Request Zip Code", "Destination Address", "Destination City",
    "Destination State", "Destination Zip Code", "Distance (miles)", "Duration (minutes)",
    "Ride Fare", "Ride Fees", "Ride Discounts", "Ride Tip", "Ride Cost", "Business Services Fee",
    "Transaction Date (UTC)", "Transaction Time (UTC)", "Transaction Amount", "Transaction Currency",
    "Transaction Type", "Expense Code", "Expense Note", "Ride Type", "Employee ID", "Custom Tag 1",
    "Custom Tag 2", "Passenger Number", "Requester Name", "Requester Email", "Internal Note",
    "Fare Type", "Scheduled Ride Id", "Flex Ride Id", "Pickup Latitude", "Pickup Longitude",
    "Drop-off Latitude", "Drop-off Longitude"
]

internal_note_values = ["FCC", "FCM", "FCSH", "FCSC", "DTF"]

def detect_header(uploaded_file):
    uploaded_file.seek(0)
    for idx in [0, 4, 5]:
        try:
            df = pd.read_csv(uploaded_file, header=idx, nrows=1)
            if any("trip/eats id" in col.lower() for col in df.columns):
                uploaded_file.seek(0)
                print( 'I found the headers on index: ', idx )
                return idx
        except Exception:
            pass
        uploaded_file.seek(0)
    return None

def clean_file_without_headers(df):
    
    # Eliminate unwanted name columns
    name_headers = ["First Name", "Last Name", "Guest First Name", "Guest Last Name"]

    if all(header in df.columns for header in name_headers):
        print( 'we have detected that there are name columns that need to be deleted' )
        df = df.drop(columns=["First Name", "Last Name"])
        df = df.rename(columns={
            "Guest First Name": "First Name",
            "Guest Last Name": "Last Name"
        })

    # Rename columns if applicable
    column_rename_map = {
        "Distance (mi)": "Distance (miles)",
        "Transaction Amount in Local Currency (incl. Taxes)": "Transaction Amount",
        "Guest Phone Number": "Passenger Number",  
        "Expense Memo": "Internal Note",           
    }
    
    df = df.rename(columns=column_rename_map)

    if 'Ride Status' in df.columns and 'Transaction Type' in df.columns:
        df['Transaction Type'] = df['Ride Status'].combine_first(df['Transaction Type'])
        df.drop(['Ride Status'], axis=1, inplace=True)
    elif 'Ride Status' in df.columns:
        df.rename(columns={"Ride Status": "Transaction Type"}, inplace=True)

    if 'Email' in df.columns and 'Requester Email' in df.columns:
        df['Email Info'] = df['Email'].combine_first(df['Requester Email'])
        df.drop(['Email', 'Requester Email'], axis=1, inplace=True)
    elif 'Email' in df.columns:
        df.rename(columns={"Email": "Email Info"}, inplace=True)
    elif 'Requester Email' in df.columns:
        df.rename(columns={"Requester Email": "Email Info"}, inplace=True)

    # Desired final columns
    desired_columns = [
        "Pickup Date (Local)",
        "Pickup Time (Local)",
        "First Name",
        "Last Name",
        "Email Info",
        "Distance (miles)",
        "Pickup Address",
        "Drop-off Address",
        "Transaction Type",
        "Internal Note",
        "Transaction Amount",
        "Passenger Number"
    ]
    
    # Keep only the desired columns that exist in the DataFrame
    final_df = df[[col for col in desired_columns if col in df.columns]].copy()
        
    return final_df

def clean_file(uploaded_file):
    try:
        print("\n📥 File received:", uploaded_file.name)
        print("📦 File type:", uploaded_file.type)
        print("📏 File size (bytes):", uploaded_file.size)

        is_common_courtesy = False
        
        if uploaded_file.name.endswith(".csv"):
            preview = pd.read_csv(uploaded_file, nrows=1, header=None)
            uploaded_file.seek(0)
            if "Common Courtesy" in str(preview.iloc[0, 1]):
                header_row = detect_header(uploaded_file)
                if header_row is not None:
                    uploaded_file.seek(0)
                    df = pd.read_csv(uploaded_file, header=header_row)
                else:
                    print("Could not detect header row — fallback to default read")
                    uploaded_file.seek(0)
                    df = pd.read_csv(uploaded_file)

                is_common_courtesy = True

                if 'Guest First Name' in df.columns:
                    df['First Name'] = df['Guest First Name']
                if 'Guest Last Name' in df.columns:
                    df['Last Name'] = df['Guest Last Name']
                df.drop(columns=['Guest First Name', 'Guest Last Name'], inplace=True, errors='ignore')
            else:
                df = pd.read_csv(uploaded_file)

            df.columns = df.columns.str.replace('\ufeff', '', regex=False).str.strip()
        else:
            df = pd.read_excel(uploaded_file)
            df.columns = df.columns.str.replace('\ufeff', '', regex=False).str.strip()

        print("📄 Header row detected at:", header_row if 'header_row' in locals() else "default (0)")
        print("🧠 Columns after cleanup:", df.columns.tolist())
        print("🧠 Are columns unique?:", df.columns.is_unique)
        print("🧪 DataFrame shape:", df.shape)

        # Eliminate unwanted name columns
        name_headers = ["First Name", "Last Name", "Guest First Name", "Guest Last Name"]

        if all(header in df.columns for header in name_headers):
            print( 'we have detected that there are name columns that need to be deleted' )
            df = df.drop(columns=["First Name", "Last Name"])
            df = df.rename(columns={
                "Guest First Name": "First Name",
                "Guest Last Name": "Last Name"
            })

        note_column = next((col for col in ['Internal Note', 'Expense Memo'] if col in df.columns), None)
        required_cols = ['First Name', 'Last Name']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if note_column is None:
            missing_cols.append('Internal Note or Expense Memo')
        if missing_cols:
            return None

        df_filtered = df[df[note_column].notna() & (df[note_column].astype(str).str.strip() != "")]
        custom_columns_to_hide = columns_to_hide.copy()
        if is_common_courtesy and "Email" in custom_columns_to_hide:
            custom_columns_to_hide.remove("Email")
        df_filtered = df_filtered.drop(columns=[col for col in custom_columns_to_hide if col in df_filtered.columns])
        if is_common_courtesy and "Transaction Type" in df_filtered.columns:
            df_filtered = df_filtered.drop(columns=["Transaction Type"])

        column_rename_map = {
            "Distance (mi)": "Distance (miles)",
            "Transaction Amount in Local Currency (incl. Taxes)": "Transaction Amount",
            "Ride Status": "Transaction Type",
            "Guest Phone Number": "Passenger Number",
            "Expense Memo": "Internal Note",
            "Email": 'Email Info',
            "Requester Email": 'Email Info',
        }

        print("🧼 Columns before rename/drop:", df_filtered.columns.tolist())
        df_filtered.rename(columns=column_rename_map, inplace=True)
        df_filtered = df_filtered.loc[:, ~df_filtered.columns.duplicated()]
        print("🧼 Columns after renaming:", df_filtered.columns.tolist())

        df_filtered['First Name'] = df_filtered['First Name'].astype(str).str.strip()
        df_filtered['Last Name'] = df_filtered['Last Name'].astype(str).str.strip()
        df_filtered['Passenger Number'] = df_filtered['Passenger Number'].astype(str).str.strip()

        df_filtered_sorted = df_filtered.sort_values(
            by=['Last Name', 'First Name', 'Passenger Number'],
            key=lambda col: col.str.lower() if col.dtype == 'object' else col
        )

        all_rows = []
        df_values = df_filtered_sorted.reset_index(drop=True)
        group_rows = []
        current_key = None

        for i in range(len(df_values)):
            row = df_values.iloc[i]
            group_key = (row['Passenger Number'], row['Last Name'], row['First Name'])
            is_last_row = (i == len(df_values) - 1)

            transaction_col = next((col for col in ["Transaction Amount", "Transaction Amount in Local Currency (incl. Taxes)"] if col in df_values.columns), None)
            row = row.copy()
            row["Fares Only"] = row.get(transaction_col or "Transaction Amount", "")

            if current_key is None:
                current_key = group_key

            if group_key != current_key:
                group_df = pd.DataFrame(group_rows).reset_index(drop=True)
                group_df["Trips Count"] = 1
                group_df["Trips Count"] = group_df["Trips Count"].astype(int)
                all_rows.extend(group_df.to_dict(orient="records"))

                total_transaction = pd.to_numeric(group_df[transaction_col], errors="coerce").sum() if transaction_col else 0
                totals_row = {col: "" for col in group_df.columns}
                totals_row[transaction_col or "Transaction Amount"] = round(total_transaction, 2)
                totals_row["Trips Count"] = int(group_df["Trips Count"].sum())
                all_rows.append(totals_row)
                all_rows.append({col: "" for col in group_df.columns})

                group_rows = []

            group_rows.append(row)
            current_key = group_key

            if is_last_row:
                group_df = pd.DataFrame(group_rows)
                group_df["Trips Count"] = 1
                group_df["Trips Count"] = group_df["Trips Count"].astype(int)
                all_rows.extend(group_df.to_dict(orient="records"))

                total_transaction = pd.to_numeric(group_df[transaction_col], errors="coerce").sum() if transaction_col else 0
                totals_row = {col: "" for col in group_df.columns}
                totals_row[transaction_col or "Transaction Amount"] = round(total_transaction, 2)
                totals_row["Trips Count"] = int(group_df["Trips Count"].sum())
                all_rows.append(totals_row)
                all_rows.append({col: "" for col in group_df.columns})

        final_df = pd.DataFrame(all_rows)

        # ✅ Add final grand total row for Fares Only
        fares_total = pd.to_numeric(final_df["Fares Only"], errors="coerce").sum()
        grand_total_row = {col: "" for col in final_df.columns}
        grand_total_row["Fares Only"] = round(fares_total, 2)
        final_df = pd.concat([final_df, pd.DataFrame([grand_total_row])], ignore_index=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Move 'Fares Only' to the last column
            if "Fares Only" in final_df.columns:
                final_df = final_df[[col for col in final_df.columns if col != "Fares Only"] + ["Fares Only"]]

            final_df.to_excel(writer, index=False, sheet_name="CleanedData")

            workbook = writer.book
            worksheet = writer.sheets["CleanedData"]

            # Define fills for each internal note type
            fills = {
                "FCC": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),   # Light Blue
                "FCM": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),   # Light Green
                "FCSH": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # Light Yellow
                "FCSC": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # Light Orange
                "DTF": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),   # Light Purple
                "Other": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), # Light Gray
            }

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Get the index of Internal Note column
            headers = [cell.value for cell in worksheet[1]]
            note_col_idx = headers.index("Internal Note") + 1  # openpyxl is 1-based

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                note_cell = row[note_col_idx - 1]
                # Check if it's a summary row (Trips Count is empty)
                trips_count_cell = row[headers.index("Trips Count")]
                is_summary_row = not trips_count_cell.value or str(trips_count_cell.value).strip() == ""

                note_value = str(note_cell.value).strip() if note_cell.value else ""

                if note_value and not is_summary_row:
                    fill = fills.get(note_value, fills["Other"])
                    for cell in row:
                        cell.fill = fill
                        cell.border = thin_border
                else:
                    for cell in row:
                        cell.border = thin_border  # Just apply borders, no fill


        output.seek(0)

        return final_df, output

    except Exception as e:
        print("Error:", e)
        return None

# --- Streamlit UI ---
st.set_page_config(page_title="Monthly Report Tool", layout="centered")
st.title("📊 Monthly Report Tool")

# --- County highlighter ---
county = st.selectbox("County highlight", ["None", "Fulton", "Forsyth"], index=0)

def highlight_rows(df, county_choice):
    if county_choice == "None" or df is None or "Internal Note" not in df.columns:
        return df
    note_sets = {
        "Fulton": {"FCC", "FCM", "FCSH", "FCSC"},
        "Forsyth": {"DTF", "DTFCE"},
    }
    target = note_sets.get(county_choice, set())

    def _row_style(row):
        note = str(row.get("Internal Note", "")).strip().upper()
        if note in target:
            # light yellow highlight
            return ["background-color: #FFF9C4"] * len(row)
        return [""] * len(row)

    try:
        return df.style.apply(_row_style, axis=1)
    except Exception:
        # If Styler not supported in the current Streamlit version, fall back to raw df
        return df

st.markdown("Upload your Excel or CSV file to clean and summarize your data.")

uploaded_file = st.file_uploader("Upload .xlsx or .csv file", type=["xlsx", "csv"])

if uploaded_file:
    cleaned_df, output = clean_file(uploaded_file)
    print("🧪 Shape after filtering:", cleaned_df)

    if cleaned_df is not None:
        st.success("✅ File cleaned successfully!")
        st.dataframe(cleaned_df.head(50))

        st.download_button(
            "📥 Download Cleaned File",
            output,
            file_name="cleaned_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

def sort_and_merge(file1_path, file2_path):
    import pandas as pd

    internal_note_values = ["FCC", "FCM", "FCSH", "FCSC", "DTF"]

    def clean_and_sort(file_obj):
        file_obj.seek(0)
        df = None  # always define df

        if file_obj.name.endswith(".csv"):
            preview = pd.read_csv(file_obj, nrows=1, header=None)
            file_obj.seek(0)

            if "Common Courtesy" in str(preview.iloc[0, 1]):
                df = pd.read_csv(file_obj, header=4)
            else:
                df = pd.read_csv(file_obj)

                if not any(col in df.columns for col in ["Last Name", "Passenger Number", "Ride ID"]):
                    file_obj.seek(0)
                    df = pd.read_csv(file_obj, header=None)

                    first_value = str(df.iloc[0, 6])
                    print("this is first value:", first_value)

                    if not any(char.isdigit() for char in first_value):
                        print("uber sheet")
                        df.columns = expected_headers_uber
                    else:
                        print("lyft sheet")
                        df.columns = expected_headers_lyft

                    df = clean_file_without_headers(df)
                else:
                    df = clean_file_without_headers(df)

        elif file_obj.name.endswith(".xlsx"):
            from openpyxl import load_workbook
            import tempfile

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(file_obj.read())
                tmp_path = tmp.name

            preview = pd.read_excel(tmp_path, nrows=5, header=None)
            if "Common Courtesy" in str(preview.iloc[0, 1]):
                df = pd.read_excel(tmp_path, header=4)
            else:
                df = pd.read_excel(tmp_path)

                if not any(col in df.columns for col in ["Last Name", "Passenger Number", "Ride ID"]):
                    df = pd.read_excel(tmp_path, header=None)

                    first_value = str(df.iloc[0, 6])
                    print("this is first value:", first_value)

                    if not any(char.isdigit() for char in first_value):
                        print("uber sheet")
                        df.columns = expected_headers_uber
                    else:
                        print("lyft sheet")
                        df.columns = expected_headers_lyft

                    df = clean_file_without_headers(df)
                else:
                    df = clean_file_without_headers(df)

        else:
            raise ValueError("Unsupported file format")

        if df is None:
            raise ValueError("Unable to load or clean the file.")

        # Clean and sort
        df['Last Name'] = df['Last Name'].astype(str).str.strip()
        df['First Name'] = df['First Name'].astype(str).str.strip()
        df['Passenger Number'] = df['Passenger Number'].astype(str).str.strip()

        if 'Internal Note' in df.columns:
            df = df[df['Internal Note'].notna() & (df['Internal Note'].astype(str).str.strip() != "")]

        return df.sort_values(
            by=['Last Name', 'First Name', 'Passenger Number'],
            key=lambda col: col.str.lower() if col.dtype == 'object' else col
        )

    df1 = clean_and_sort(file1_path)
    df2 = clean_and_sort(file2_path)

    combined_df = pd.concat([df1, df2], ignore_index=True)
    df_sorted = combined_df.sort_values(
        by=['Last Name', 'First Name', 'Passenger Number'],
        key=lambda col: col.str.lower() if col.dtype == 'object' else col
    ).reset_index(drop=True)

    all_rows = []
    df_values = df_sorted.reset_index(drop=True)
    group_rows = []
    current_key = None

    for i in range(len(df_sorted)):
        row = df_sorted.iloc[i]
        group_key = (row['Passenger Number'], row['Last Name'], row['First Name'])
        is_last_row = (i == len(df_sorted) - 1)

        transaction_col = next((col for col in ["Transaction Amount", "Transaction Amount in Local Currency (incl. Taxes)"] if col in df_values.columns), None)
        row = row.copy()
        row["Fares Only"] = row.get(transaction_col or "Transaction Amount", "")

        if current_key is None:
            current_key = group_key

        if group_key != current_key:
            group_df = pd.DataFrame(group_rows)
            group_df["Trips Count"] = 1
            group_df["Trips Count"] = group_df["Trips Count"].astype(int)
            all_rows.extend(group_df.to_dict(orient="records"))

            transaction_col = next((col for col in ["Transaction Amount", "Transaction Amount in Local Currency (incl. Taxes)"] if col in group_df.columns), None)
            total_transaction = pd.to_numeric(group_df[transaction_col], errors="coerce").sum() if transaction_col else 0

            totals_row = {col: "" for col in group_df.columns}
            totals_row[transaction_col or "Transaction Amount"] = round(total_transaction, 2)
            totals_row["Trips Count"] = int(group_df["Trips Count"].sum())
            all_rows.append(totals_row)
            all_rows.append({col: "" for col in group_df.columns})  # Empty row

            group_rows = []

        group_rows.append(row)
        current_key = group_key

        if is_last_row:
            group_df = pd.DataFrame(group_rows)
            group_df["Trips Count"] = 1
            group_df["Trips Count"] = group_df["Trips Count"].astype(int)
            all_rows.extend(group_df.to_dict(orient="records"))

            transaction_col = next((col for col in ["Transaction Amount", "Transaction Amount in Local Currency (incl. Taxes)"] if col in group_df.columns), None)
            total_transaction = pd.to_numeric(group_df[transaction_col], errors="coerce").sum() if transaction_col else 0

            totals_row = {col: "" for col in group_df.columns}
            totals_row[transaction_col or "Transaction Amount"] = round(total_transaction, 2)
            totals_row["Trips Count"] = int(group_df["Trips Count"].sum())
            all_rows.append(totals_row)
            all_rows.append({col: "" for col in group_df.columns})

    final_df = pd.DataFrame(all_rows)

    # ✅ Add final grand total row for Fares Only
    fares_total = pd.to_numeric(final_df["Fares Only"], errors="coerce").sum()
    grand_total_row = {col: "" for col in final_df.columns}
    grand_total_row["Fares Only"] = round(fares_total, 2)
    final_df = pd.concat([final_df, pd.DataFrame([grand_total_row])], ignore_index=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Move 'Fares Only' to the last column
        if "Fares Only" in final_df.columns:
            final_df = final_df[[col for col in final_df.columns if col != "Fares Only"] + ["Fares Only"]]

        final_df.to_excel(writer, index=False, sheet_name="CleanedData")
        workbook = writer.book
        worksheet = writer.sheets["CleanedData"]

        # Color fills by internal note
        fills = {
            "FCC": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
            "FCM": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "FCSH": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "FCSC": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "DTF": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),
            "Other": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        }

        # Define a thin border for all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Get column indices
        headers = [cell.value for cell in worksheet[1]]
        note_col_idx = headers.index("Internal Note") + 1 if "Internal Note" in headers else None
        trips_col_idx = headers.index("Trips Count") + 1 if "Trips Count" in headers else None

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            # Detect if row is a summary row
            is_summary_row = trips_col_idx and (not row[trips_col_idx - 1].value)

            # Highlight based on Internal Note only if not a summary row
            if note_col_idx and not is_summary_row:
                note_value = str(row[note_col_idx - 1].value).strip() if row[note_col_idx - 1].value else ""
                if note_value:
                    fill = fills.get(note_value, fills["Other"])
                    for cell in row:
                        cell.fill = fill
                        cell.border = thin_border
                    continue

            # Otherwise, apply only borders (no fill)
            for cell in row:
                cell.border = thin_border

    output.seek(0)
    return final_df, output

def split_by_internal_note(df):
    """
    Exports grouped files:
      - "Forsyth": DTF + DTFCE (with Forsyth billing columns)
      - "Fulton":  FCC + FCM + FCSH + FCSC (combined)
      - "Other_report": any other non-empty Internal Note values
    Returns: dict[str, tuple[pd.DataFrame, BytesIO]]
    """
    split_files = {}

    if 'Internal Note' not in df.columns:
        return {}

    # Normalize key fields
    df['Last Name'] = df['Last Name'].astype(str).str.strip()
    df['First Name'] = df['First Name'].astype(str).str.strip()
    df['Passenger Number'] = df['Passenger Number'].astype(str).str.strip()

    # Case-insensitive notes
    norm_notes = df['Internal Note'].astype(str).str.strip().str.upper()

    # Note groups
    forsyth_notes = {"DTF", "DTFCE"}
    fulton_notes  = {"FCC", "FCM", "FCSH", "FCSC"}

    # Split datasets
    dtf_mask = norm_notes.isin(forsyth_notes)
    df_forsyth = df[dtf_mask].copy()
    df_fulton  = df[norm_notes.isin(fulton_notes)].copy()

    # Remaining = anything not Forsyth or Fulton but still has a note
    remaining = df[~norm_notes.isin(forsyth_notes.union(fulton_notes))].copy()
    other_df = remaining[
        remaining['Internal Note'].notna() &
        (remaining['Internal Note'].astype(str).str.strip() != "")
    ]

    def group_and_export(df_note, is_dtf=False):
        all_rows = []
        group_rows = []
        current_key = None

        # prefer "Fare", fallback to "Fares Only"
        fare_candidates = [c for c in ("Fare", "Fares Only") if c in df_note.columns]
        fare_col = fare_candidates[0] if fare_candidates else None

        for i in range(len(df_note)):
            base_row = df_note.iloc[i]
            group_key = (base_row['Passenger Number'], base_row['Last Name'], base_row['First Name'])
            is_last_row = (i == len(df_note) - 1)

            if current_key is None:
                current_key = group_key

            if group_key != current_key:
                group_df = pd.DataFrame(group_rows)
                group_df["Trips Count"] = 1

                # per-user totals on last row
                if is_dtf and len(group_df):
                    # TOTAL Forsyth Bill
                    if "Forsyth Bill" in group_df.columns:
                        group_df["TOTAL Forsyth Bill"] = None
                        total_fb = pd.to_numeric(group_df["Forsyth Bill"], errors="coerce").sum()
                        group_df.at[group_df.index[-1], "TOTAL Forsyth Bill"] = round(total_fb, 2)
                    # TOTAL Rider Cost Rider Bill
                    if "Rider Cost Rider Bill" in group_df.columns:
                        group_df["TOTAL Rider Cost Rider Bill"] = None
                        total_rcb = pd.to_numeric(group_df["Rider Cost Rider Bill"], errors="coerce").sum()
                        group_df.at[group_df.index[-1], "TOTAL Rider Cost Rider Bill"] = round(total_rcb, 2)

                all_rows.extend(group_df.to_dict(orient="records"))

                transaction_col = next(
                    (col for col in ["Transaction Amount", "Transaction Amount in Local Currency (incl. Taxes)"]
                    if col in group_df.columns),
                    None
                )
                total_transaction = pd.to_numeric(group_df[transaction_col], errors="coerce").sum() if transaction_col else 0

                totals_row = {col: None for col in group_df.columns}
                totals_row[transaction_col or "Transaction Amount"] = round(total_transaction, 2)
                totals_row["Trips Count"] = int(pd.to_numeric(group_df["Trips Count"], errors="coerce").sum())
                all_rows.append(totals_row)

                spacer_row = {col: None for col in group_df.columns}
                all_rows.append(spacer_row)

                group_rows = []

            # ---- user row ----
            row = base_row.copy()

            if is_dtf:
                # Rider Co-Pay = 5.00
                row["Rider Co-Pay"] = 5.00

                # Post Co-Pay Cost = Fare - Rider Co-Pay
                post = None
                fare_val = None
                if fare_col is not None:
                    fare_val = pd.to_numeric(row.get(fare_col), errors="coerce")
                    if pd.notna(fare_val):
                        post = round(fare_val - 5.00, 2)
                row["Post Co-Pay Cost"] = post

                # Forsyth Bill = MIN(8, MAX(0, Fare - 5))
                forsyth_bill = None
                if post is not None:
                    forsyth_bill = round(min(8.00, max(0.00, post)), 2)
                row["Forsyth Bill"] = forsyth_bill

                # Rider Share over $13 = IF(Fare > 13, Fare - 13, 0)
                share_over_13 = None
                if fare_val is not None and pd.notna(fare_val):
                    share_over_13 = round(max(0.00, fare_val - 13.00), 2)
                row["Rider Share over $13"] = share_over_13

                # Rider Cost Rider Bill = Rider Co-Pay + Rider Share over $13
                rider_cost = None
                rc = pd.to_numeric(row.get("Rider Co-Pay"), errors="coerce")
                r13 = pd.to_numeric(row.get("Rider Share over $13"), errors="coerce")
                if pd.notna(rc) and pd.notna(r13):
                    rider_cost = round(rc + r13, 2)
                row["Rider Cost Rider Bill"] = rider_cost

            group_rows.append(row)
            current_key = group_key

            if is_last_row:
                group_df = pd.DataFrame(group_rows)
                group_df["Trips Count"] = 1

                # per-user totals on last row (final group)
                if is_dtf and len(group_df):
                    if "Forsyth Bill" in group_df.columns:
                        group_df["TOTAL Forsyth Bill"] = None
                        total_fb = pd.to_numeric(group_df["Forsyth Bill"], errors="coerce").sum()
                        group_df.at[group_df.index[-1], "TOTAL Forsyth Bill"] = round(total_fb, 2)
                    if "Rider Cost Rider Bill" in group_df.columns:
                        group_df["TOTAL Rider Cost Rider Bill"] = None
                        total_rcb = pd.to_numeric(group_df["Rider Cost Rider Bill"], errors="coerce").sum()
                        group_df.at[group_df.index[-1], "TOTAL Rider Cost Rider Bill"] = round(total_rcb, 2)

                all_rows.extend(group_df.to_dict(orient="records"))

                transaction_col = next(
                    (col for col in ["Transaction Amount", "Transaction Amount in Local Currency (incl. Taxes)"]
                    if col in group_df.columns),
                    None
                )
                total_transaction = pd.to_numeric(group_df[transaction_col], errors="coerce").sum() if transaction_col else 0

                totals_row = {col: None for col in group_df.columns}
                totals_row[transaction_col or "Transaction Amount"] = round(total_transaction, 2)
                totals_row["Trips Count"] = int(pd.to_numeric(group_df["Trips Count"], errors="coerce").sum())
                all_rows.append(totals_row)

                spacer_row = {col: None for col in group_df.columns}
                all_rows.append(spacer_row)

        final_df = pd.DataFrame(all_rows)

        # Drop unwanted columns in final output
        drop_cols = [
            "Transaction Type",
            "Transaction Amount",
            "Passenger Number",
            "Email Info",
            "Trips Count",
        ]
        final_df = final_df.drop(columns=[c for c in drop_cols if c in final_df.columns], errors="ignore")

        # Ensure numeric dtypes
        for col in [
            "Transaction Amount",
            "Transaction Amount in Local Currency (incl. Taxes)",
            "Fare",
            "Fares Only",
            "Trips Count",
            "Rider Co-Pay",
            "Post Co-Pay Cost",
            "Forsyth Bill",
            "Rider Share over $13",
            "Rider Cost Rider Bill",
            "TOTAL Forsyth Bill",
            "TOTAL Rider Cost Rider Bill",
        ]:
            if col in final_df.columns:
                final_df[col] = pd.to_numeric(final_df[col], errors="coerce")

        # Put Forsyth fields at the end (only when is_dtf)
        if is_dtf:
            tail_cols = [
                "Rider Co-Pay",
                "Post Co-Pay Cost",
                "Forsyth Bill",
                "Rider Share over $13",
                "Rider Cost Rider Bill",
                "TOTAL Rider Cost Rider Bill",
                "TOTAL Forsyth Bill",
            ]
            for tail_col in tail_cols:
                if tail_col not in final_df.columns:
                    final_df[tail_col] = None
            cols = [c for c in final_df.columns if c not in tail_cols]
            cols.extend(tail_cols)
            final_df = final_df[cols]

        # Reorder the last 3 columns exactly as requested
        end_order = ["Internal Note", "TOTAL Forsyth Bill", "TOTAL Rider Cost Rider Bill"]
        present = [c for c in end_order if c in final_df.columns]
        other_cols = [c for c in final_df.columns if c not in present]
        final_df = final_df[other_cols + present]

        # Grand total for Fare (or Fares Only)
        fare_total_col = "Fare" if "Fare" in final_df.columns else ("Fares Only" if "Fares Only" in final_df.columns else None)
        if fare_total_col:
            fares_total = pd.to_numeric(final_df[fare_total_col], errors="coerce").sum()
            grand_total_row = {col: None for col in final_df.columns}
            grand_total_row[fare_total_col] = round(fares_total, 2)
            final_df = pd.concat([final_df, pd.DataFrame([grand_total_row])], ignore_index=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_name = "Sheet1"
            final_df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            # Currency formatting
            currency_cols = [
                fare_total_col,
                "Rider Co-Pay",
                "Post Co-Pay Cost",
                "Forsyth Bill",
                "Rider Share over $13",
                "Rider Cost Rider Bill",
                "TOTAL Rider Cost Rider Bill",
                "TOTAL Forsyth Bill",
            ]
            
            for col_name in [c for c in currency_cols if c and c in final_df.columns]:
                col_idx = final_df.columns.get_loc(col_name) + 1
                for r in range(2, len(final_df) + 2):
                    ws.cell(row=r, column=col_idx).number_format = '"$"#,##0.00'


        output.seek(0)
        return final_df, output

    # --- Exports ---

    # DTF combined (DTF + DTFCE)
    if not df_forsyth.empty:
        split_files["Forsyth"] = group_and_export(df_forsyth, is_dtf=True)

    if not df_fulton.empty:
        split_files["Fulton"] = group_and_export(df_fulton, is_dtf=False)

    if not other_df.empty:
        split_files["Other_report"] = group_and_export(other_df, is_dtf=False)

    return split_files


# --- Combine Two Files Section (UI) ---
st.markdown("#### 📎 Combine Two Filtered Files & Split by Internal Notes")

# Placeholder button directly under header (disabled until merge)
btn_slot = st.empty()
btn_slot.download_button(
    label="📥 Download Merged File",
    data=b"",  # placeholder
    file_name="merged_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    disabled=True,
    help="Upload both files to enable"
)

# Uploaders live under the (currently disabled) button
uploaded_file1 = st.file_uploader("Upload your first .xlsx or .csv file", type=["xlsx", "csv"], key="file1")
uploaded_file2 = st.file_uploader("Upload your second .xlsx or .csv file", type=["xlsx", "csv"], key="file2")

if uploaded_file1 and uploaded_file2:
    try:
        with st.spinner("🟡 Merging files..."):
            df, output = sort_and_merge(uploaded_file1, uploaded_file2)

        st.success("✅ Files merged successfully!")
        st.dataframe(df.head(50))

        # Swap disabled button for a real download button
        btn_slot.download_button(
            label="📥 Download Merged File",
            data=output,
            file_name="merged_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            disabled=False
        )

        # Split by note
        st.markdown("### 🔍 Download Split Files by Internal Note")
        split_files = split_by_internal_note(df)
        if not split_files:
            st.warning("⚠️ Could not find 'Internal Note' column to split by.")
        else:
            if county == "Forsyth":
                keys_to_show = ["Forsyth"]
            elif county == "Fulton":
                keys_to_show = ["Fulton"]
            else:
                keys_to_show = list(split_files.keys())

            shown = False
            for key in keys_to_show:
                if key in split_files:
                    df_note, buffer = split_files[key]
                    st.download_button(
                        label=f"📂 Download {key}.xlsx",
                        data=buffer,
                        file_name=f"{key}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    shown = True
            if not shown:
                st.info("No files found for the selected county.")
    except Exception as e:
        st.error(f"❌ Merge failed: {e}")
